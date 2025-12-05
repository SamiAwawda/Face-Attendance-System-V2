"""
AI-Powered Face Attendance System
Flask Application with Firebase Integration
"""

from flask import Flask, render_template, Response, request, jsonify, session, redirect, url_for
from camera_engine import get_engine, CameraEngine
from datetime import datetime, timedelta
from threading import Lock
from functools import wraps
import firebase_admin
from firebase_admin import credentials, firestore
import os
import json

# Initialize Flask
app = Flask(__name__)
app.secret_key = os.urandom(24)

# Admin password (simple authentication)
ADMIN_PASSWORD = "admin123"


def login_required(f):
    """Decorator to require admin login for protected routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Firebase initialization
db = None
firebase_initialized = False

def init_firebase():
    """Initialize Firebase connection."""
    global db, firebase_initialized
    
    if firebase_initialized:
        return True
    
    try:
        # Check for credentials file
        cred_path = os.path.join(os.path.dirname(__file__), 'firebase_config.json')
        
        if not os.path.exists(cred_path):
            print("Warning: firebase_config.json not found. Using mock database.")
            return False
        
        cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred)
        db = firestore.client()
        firebase_initialized = True
        print("Firebase initialized successfully")
        return True
    except Exception as e:
        print(f"Firebase initialization error: {e}")
        return False

# Initialize Firebase on startup
init_firebase()

# In-memory cache for recent logs (session-based)
recent_logs = []
recent_logs_lock = Lock()


def load_known_faces():
    """Load all registered student embeddings from Firebase."""
    engine = get_engine()
    
    if not firebase_initialized or db is None:
        print("Firebase not available, cannot load faces")
        return False
    
    try:
        students_ref = db.collection('students')
        docs = students_ref.stream()
        
        embeddings_list = []
        for doc in docs:
            try:
                data = doc.to_dict()
                if data and 'embedding' in data and 'student_id' in data and 'name' in data:
                    embeddings_list.append((
                        data['student_id'],
                        data['name'],
                        data['embedding']
                    ))
            except Exception as doc_error:
                print(f"Warning: Error processing student document: {doc_error}")
                continue
        
        engine.load_embeddings(embeddings_list)
        print(f"Loaded {len(embeddings_list)} student faces from Firebase")
        return True
    except Exception as e:
        print(f"Error loading faces: {e}")
        import traceback
        traceback.print_exc()
        return False


def check_and_mark_attendance(student_id, name):
    """
    Check if attendance can be marked (60-minute rule) and mark if allowed.
    Uses a simpler query to avoid Firestore composite index requirement.
    
    Returns:
        (status, color) tuple
    """
    global recent_logs
    
    now = datetime.now()
    one_hour_ago = now - timedelta(minutes=60)
    
    if firebase_initialized and db is not None:
        try:
            # Simpler query: get recent attendance for this student (no composite index needed)
            attendance_ref = db.collection('attendance')
            
            # Query by student_id only, then filter by time in Python
            docs = attendance_ref.where('student_id', '==', student_id).stream()
            
            # Check if any entry is within the last 60 minutes
            for doc in docs:
                data = doc.to_dict()
                last_time = data.get('timestamp')
                
                # Handle Firestore timestamp
                if hasattr(last_time, 'timestamp'):
                    last_time = datetime.fromtimestamp(last_time.timestamp())
                elif isinstance(last_time, str):
                    last_time = datetime.fromisoformat(last_time.replace('Z', '+00:00'))
                elif not isinstance(last_time, datetime):
                    continue
                
                # Check 60-minute rule
                if last_time > one_hour_ago:
                    print(f"Already marked: {name} at {last_time}")
                    return "already_marked", CameraEngine.COLOR_GOLD
            
            # Mark new attendance
            attendance_ref.add({
                'student_id': student_id,
                'name': name,
                'timestamp': now,
                'status': 'success'
            })
            
            print(f"Attendance marked: {name} at {now.strftime('%H:%M:%S')}")
            
            # Add to recent logs cache
            with recent_logs_lock:
                recent_logs.insert(0, {
                    'student_id': student_id,
                    'name': name,
                    'time': now.strftime('%H:%M:%S'),
                    'status': 'success'
                })
                # Keep only last 20 entries
                recent_logs = recent_logs[:20]
            
            return "success", CameraEngine.COLOR_GREEN
            
        except Exception as e:
            print(f"Attendance error: {e}")
            import traceback
            traceback.print_exc()
            return "error", CameraEngine.COLOR_RED
    
    # Fallback for when Firebase is not available
    return "success", CameraEngine.COLOR_GREEN


# Set up attendance callback
def setup_engine_callback():
    """Configure the camera engine with attendance callback."""
    engine = get_engine()
    engine.on_attendance_marked = check_and_mark_attendance


# ============================================================
# ROUTES
# ============================================================

@app.route('/')
def index():
    """Landing page with mode selection."""
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page."""
    if session.get('admin_logged_in'):
        return redirect(url_for('register'))
    
    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('register'))
        else:
            error = 'Invalid password. Please try again.'
    
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    """Logout and clear session."""
    session.pop('admin_logged_in', None)
    return redirect(url_for('index'))


@app.route('/register')
@login_required
def register():
    """Student Registration Page (requires login)."""
    engine = get_engine()
    if not engine.is_running:
        engine.start()
    
    # Load known faces for duplicate detection
    try:
        load_known_faces()
    except Exception as e:
        print(f"Warning: Could not load faces for registration: {e}")
    
    return render_template('register.html')


@app.route('/attendance')
def attendance():
    """Attendance kiosk page."""
    global recent_logs
    
    try:
        engine = get_engine()
        if not engine.is_running:
            engine.start()
        
        # Load known faces for recognition (with error handling)
        try:
            load_known_faces()
        except Exception as e:
            print(f"Warning: Could not load faces, continuing without recognition: {e}")
        
        # Setup callback (wrapped in try-except)
        try:
            setup_engine_callback()
        except Exception as e:
            print(f"Warning: Could not setup attendance callback: {e}")
        
        # Clear previous session logs
        with recent_logs_lock:
            recent_logs = []
        
        return render_template('attendance.html')
    except Exception as e:
        print(f"Critical error in attendance route: {e}")
        import traceback
        traceback.print_exc()
        return render_template('attendance.html')


@app.route('/video_feed/<mode>')
def video_feed(mode):
    """Stream video feed with bounding boxes."""
    engine = get_engine()
    if not engine.is_running:
        engine.start()
    
    def generate():
        while True:
            frame = engine.get_frame(mode=mode)
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
    
    return Response(
        generate(),
        mimetype='multipart/x-mixed-replace; boundary=frame'
    )


@app.route('/api/register', methods=['POST'])
def api_register():
    """Register a new student with face capture."""
    data = request.get_json()
    
    if not data:
        return jsonify({'success': False, 'message': 'No data provided'}), 400
    
    student_id = data.get('student_id', '').strip()
    name = data.get('name', '').strip()
    
    if not student_id or not name:
        return jsonify({'success': False, 'message': 'Student ID and Name are required'}), 400
    
    # Capture face and get embedding
    engine = get_engine()
    result = engine.capture_face()
    
    if not result['success']:
        return jsonify(result), 400
    
    embedding = result['embedding']
    
    # Save to Firebase
    if firebase_initialized and db is not None:
        try:
            # Check if student already exists
            students_ref = db.collection('students')
            existing = students_ref.where('student_id', '==', student_id).limit(1).stream()
            
            if list(existing):
                return jsonify({
                    'success': False,
                    'message': f'Student ID {student_id} is already registered'
                }), 400
            
            # Add new student
            students_ref.add({
                'student_id': student_id,
                'name': name,
                'embedding': embedding,
                'created_at': datetime.now()
            })
            
            # Reload embeddings
            load_known_faces()
            
            return jsonify({
                'success': True,
                'message': f'Successfully registered {name} (ID: {student_id})'
            })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Database error: {str(e)}'
            }), 500
    else:
        return jsonify({
            'success': False,
            'message': 'Database not available. Please configure firebase_config.json'
        }), 500


@app.route('/api/face_status')
def api_face_status():
    """Get current face status for registration page (duplicate prevention)."""
    try:
        engine = get_engine()
        
        with engine.face_results_lock:
            # Get the most recent face result
            if engine.face_results:
                # Get the latest result by timestamp
                latest_key = None
                latest_time = None
                for key, (color, name, status, timestamp) in engine.face_results.items():
                    if latest_time is None or timestamp > latest_time:
                        latest_time = timestamp
                        latest_key = key
                
                if latest_key:
                    color, name, status, timestamp = engine.face_results[latest_key]
                    return jsonify({
                        'success': True,
                        'has_face': True,
                        'status': status,  # 'registered', 'new', or 'scanning'
                        'name': name if status == 'registered' else None,
                        'can_register': status == 'new'
                    })
        
        return jsonify({
            'success': True,
            'has_face': False,
            'status': 'no_face',
            'name': None,
            'can_register': False
        })
    except Exception as e:
        print(f"Face status error: {e}")
        return jsonify({
            'success': False,
            'has_face': False,
            'status': 'error',
            'name': None,
            'can_register': True  # Allow registration on error
        })


@app.route('/api/recent_logs')
def api_recent_logs():
    """Get recent attendance entries for sidebar."""
    with recent_logs_lock:
        return jsonify({
            'success': True,
            'logs': recent_logs[:15]  # Return last 15 entries
        })


@app.route('/api/recent_registrations')
def api_recent_registrations():
    """Get recently registered students from Firebase (persists across restarts)."""
    if not firebase_initialized or db is None:
        return jsonify({
            'success': False,
            'message': 'Database not available',
            'registrations': []
        })
    
    try:
        students_ref = db.collection('students')
        
        # Query students ordered by created_at (descending) and limit to 20
        docs = students_ref.order_by('created_at', direction=firestore.Query.DESCENDING) \
                          .limit(20) \
                          .stream()
        
        registrations = []
        for doc in docs:
            data = doc.to_dict()
            created_at = data.get('created_at')
            
            # Handle Firestore timestamp
            if hasattr(created_at, 'timestamp'):
                time_str = datetime.fromtimestamp(created_at.timestamp()).strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(created_at, datetime):
                time_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                time_str = str(created_at) if created_at else 'N/A'
            
            registrations.append({
                'student_id': data.get('student_id', 'N/A'),
                'name': data.get('name', 'Unknown'),
                'registered_at': time_str
            })
        
        return jsonify({
            'success': True,
            'registrations': registrations,
            'count': len(registrations)
        })
        
    except Exception as e:
        print(f"Recent registrations error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': str(e),
            'registrations': []
        })


@app.route('/api/stats')
def api_stats():
    """Get attendance statistics."""
    if not firebase_initialized or db is None:
        return jsonify({
            'success': False,
            'message': 'Database not available'
        })
    
    try:
        # Get student count
        students = list(db.collection('students').stream())
        student_count = len(students)
        
        # Get today's attendance count
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        attendance_ref = db.collection('attendance')
        today_attendance = list(
            attendance_ref.where('timestamp', '>=', today_start).stream()
        )
        
        return jsonify({
            'success': True,
            'stats': {
                'total_students': student_count,
                'today_entries': len(today_attendance)
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        })


@app.route('/admin')
@login_required
def admin_dashboard():
    """Admin dashboard with daily attendance log (requires login)."""
    today = datetime.now().strftime('%A, %B %d, %Y')
    return render_template('admin.html', today_date=today)


@app.route('/api/daily_logs')
def api_daily_logs():
    """Get today's attendance logs for admin dashboard."""
    if not firebase_initialized or db is None:
        return jsonify({
            'success': False,
            'message': 'Database not available',
            'logs': []
        })
    
    try:
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        attendance_ref = db.collection('attendance')
        
        # Query today's attendance ordered by time
        docs = attendance_ref.where('timestamp', '>=', today_start) \
                            .order_by('timestamp', direction=firestore.Query.DESCENDING) \
                            .stream()
        
        logs = []
        for doc in docs:
            data = doc.to_dict()
            timestamp = data.get('timestamp')
            
            # Handle Firestore timestamp
            if hasattr(timestamp, 'timestamp'):
                time_str = datetime.fromtimestamp(timestamp.timestamp()).strftime('%H:%M:%S')
            elif isinstance(timestamp, datetime):
                time_str = timestamp.strftime('%H:%M:%S')
            else:
                time_str = str(timestamp)
            
            logs.append({
                'name': data.get('name', 'Unknown'),
                'student_id': data.get('student_id', 'N/A'),
                'time': time_str,
                'status': data.get('status', 'success')
            })
        
        return jsonify({
            'success': True,
            'logs': logs,
            'count': len(logs),
            'date': datetime.now().strftime('%Y-%m-%d')
        })
    except Exception as e:
        print(f"Daily logs error: {e}")
        return jsonify({
            'success': False,
            'message': str(e),
            'logs': []
        })


@app.route('/api/export_csv')
def api_export_csv():
    """Export today's attendance as CSV."""
    if not firebase_initialized or db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    
    try:
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        attendance_ref = db.collection('attendance')
        docs = attendance_ref.where('timestamp', '>=', today_start) \
                            .order_by('timestamp') \
                            .stream()
        
        # Build CSV content
        csv_lines = ['Student Name,Student ID,Time of Entry']
        
        for doc in docs:
            data = doc.to_dict()
            timestamp = data.get('timestamp')
            
            if hasattr(timestamp, 'timestamp'):
                time_str = datetime.fromtimestamp(timestamp.timestamp()).strftime('%H:%M:%S')
            elif isinstance(timestamp, datetime):
                time_str = timestamp.strftime('%H:%M:%S')
            else:
                time_str = str(timestamp)
            
            name = data.get('name', 'Unknown').replace(',', ' ')
            student_id = data.get('student_id', 'N/A')
            csv_lines.append(f'{name},{student_id},{time_str}')
        
        csv_content = '\n'.join(csv_lines)
        
        response = Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=attendance_{today_str}.csv'
            }
        )
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/export_students_excel')
@login_required
def api_export_students_excel():
    """Export all registered students to Excel file."""
    if not firebase_initialized or db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registered Students"
        
        # Header row with styling
        headers = ['#', 'Student Name', 'Student ID', 'Registration Date']
        ws.append(headers)
        
        # Style headers
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
        
        # Fetch all students
        students_ref = db.collection('students')
        docs = students_ref.order_by('created_at', direction=firestore.Query.DESCENDING).stream()
        
        row_num = 1
        for doc in docs:
            data = doc.to_dict()
            created_at = data.get('created_at')
            
            # Handle Firestore timestamp
            if hasattr(created_at, 'timestamp'):
                time_str = datetime.fromtimestamp(created_at.timestamp()).strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(created_at, datetime):
                time_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                time_str = str(created_at) if created_at else 'N/A'
            
            ws.append([
                row_num,
                data.get('name', 'Unknown'),
                data.get('student_id', 'N/A'),
                time_str
            ])
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 22
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        today_str = datetime.now().strftime('%Y-%m-%d')
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=students_{today_str}.xlsx'
            }
        )
        
    except ImportError:
        return jsonify({
            'success': False, 
            'message': 'openpyxl library not installed. Run: pip install openpyxl'
        }), 500
    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/delete_student/<student_id>', methods=['DELETE'])
@login_required
def api_delete_student(student_id):
    """Delete a student from the database."""
    if not firebase_initialized or db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    
    try:
        students_ref = db.collection('students')
        
        # Find the student document by student_id
        docs = students_ref.where('student_id', '==', student_id).stream()
        
        deleted = False
        for doc in docs:
            doc.reference.delete()
            deleted = True
        
        if deleted:
            # Reload known faces to remove from cache
            try:
                load_known_faces()
            except Exception as e:
                print(f"Warning: Could not reload faces after deletion: {e}")
            
            return jsonify({
                'success': True,
                'message': f'Student {student_id} deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Student {student_id} not found'
            }), 404
            
    except Exception as e:
        print(f"Delete error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/all_students')
@login_required
def api_all_students():
    """Get all registered students for admin panel."""
    if not firebase_initialized or db is None:
        return jsonify({
            'success': False,
            'message': 'Database not available',
            'students': []
        })
    
    try:
        students_ref = db.collection('students')
        docs = students_ref.order_by('created_at', direction=firestore.Query.DESCENDING).stream()
        
        students = []
        for doc in docs:
            data = doc.to_dict()
            created_at = data.get('created_at')
            
            # Handle Firestore timestamp
            if hasattr(created_at, 'timestamp'):
                time_str = datetime.fromtimestamp(created_at.timestamp()).strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(created_at, datetime):
                time_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                time_str = str(created_at) if created_at else 'N/A'
            
            students.append({
                'student_id': data.get('student_id', 'N/A'),
                'name': data.get('name', 'Unknown'),
                'registered_at': time_str
            })
        
        return jsonify({
            'success': True,
            'students': students,
            'count': len(students)
        })
        
    except Exception as e:
        print(f"All students error: {e}")
        return jsonify({
            'success': False,
            'message': str(e),
            'students': []
        })


@app.route('/api/attendance_by_date')
def api_attendance_by_date():
    """Get all attendance records grouped by date."""
    if not firebase_initialized or db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    
    try:
        attendance_ref = db.collection('attendance')
        docs = attendance_ref.order_by('timestamp', direction=firestore.Query.DESCENDING).stream()
        
        # Group by date
        dates = {}
        
        for doc in docs:
            data = doc.to_dict()
            timestamp = data.get('timestamp')
            
            # Get date and time from timestamp
            if hasattr(timestamp, 'timestamp'):
                dt = datetime.fromtimestamp(timestamp.timestamp())
            elif isinstance(timestamp, datetime):
                dt = timestamp
            else:
                continue  # Skip if no valid timestamp
            
            date_str = dt.strftime('%Y-%m-%d')
            time_str = dt.strftime('%H:%M:%S')
            
            if date_str not in dates:
                dates[date_str] = []
            
            dates[date_str].append({
                'name': data.get('name', 'Unknown'),
                'student_id': data.get('student_id', 'N/A'),
                'time': time_str
            })
        
        return jsonify({
            'success': True,
            'dates': dates
        })
        
    except Exception as e:
        print(f"Attendance by date error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/export_attendance_excel')
def api_export_attendance_excel():
    """Export attendance for a specific date to Excel file."""
    if not firebase_initialized or db is None:
        return jsonify({'success': False, 'message': 'Database not available'}), 500
    
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        # Get date parameter
        date_str = request.args.get('date')
        if not date_str:
            date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Parse the date
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d')
            next_date = target_date + timedelta(days=1)
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {date_str}"
        
        # Header row
        headers = ['#', 'Student Name', 'Student ID', 'Time']
        ws.append(headers)
        
        # Style headers
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
        
        # Query attendance for that day
        attendance_ref = db.collection('attendance')
        docs = attendance_ref.where(
            'timestamp', '>=', target_date
        ).where(
            'timestamp', '<', next_date
        ).order_by('timestamp').stream()
        
        row_num = 1
        for doc in docs:
            data = doc.to_dict()
            timestamp = data.get('timestamp')
            
            if hasattr(timestamp, 'timestamp'):
                time_str = datetime.fromtimestamp(timestamp.timestamp()).strftime('%H:%M:%S')
            elif isinstance(timestamp, datetime):
                time_str = timestamp.strftime('%H:%M:%S')
            else:
                time_str = 'N/A'
            
            ws.append([
                row_num,
                data.get('name', 'Unknown'),
                data.get('student_id', 'N/A'),
                time_str
            ])
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=attendance_{date_str}.xlsx'
            }
        )
        
    except ImportError:
        return jsonify({
            'success': False, 
            'message': 'openpyxl library not installed'
        }), 500
    except Exception as e:
        print(f"Export attendance error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    print("=" * 50)
    print("AI-Powered Face Attendance System")
    print("=" * 50)
    print("\nStarting Flask server...")
    print("Open http://localhost:5000 in your browser\n")
    
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
